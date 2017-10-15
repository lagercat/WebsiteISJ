import os

import django
import geocoder
import unicodedata

from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from school.models import School

wb = load_workbook('retea.xlsx')
count = 0
total_count = 0

to_eliminate = ['MUN', 'COM', 'ORAS']

for sheet in wb.get_sheet_names():
    for row in range(2, wb[sheet].max_row):
        eliminating_word = ''
        name_file_row = 'B' + str(row)
        address_file_row = 'C' + str(row)
        school_name = wb[sheet][name_file_row].value
        school_adress = wb[sheet][address_file_row].value
        address_partitioned = school_adress.split(",")
        address_partitioned.pop(0)
        city = address_partitioned[0]
        city = unicodedata.normalize("NFKD", city).encode("ascii", "ignore")
        address_partitioned.pop(len(address_partitioned) - 1)
        for i in range(0, len(address_partitioned) - 1):
            address_partitioned[i] += " "
        school_adress = "".join(address_partitioned)
        school_location = geocoder.google(school_adress)
        if (not school_location.latlng and "timisoara" not in city.lower()) or "principala" in school_adress.lower():
            school_location = geocoder.google(city)
        if school_location.latlng is not None:
            school_coordinates = str(school_location.latlng[0]) + "," + str(school_location.latlng[1])
            School.objects.create(name=school_name,
                                  address=school_location.address,
                                  geolocation=school_coordinates)
            count += 1
        else:
            print school_adress
        total_count += 1
    print "Count is " + str(count) + "\n"
    print "Total count is " + str(total_count)
