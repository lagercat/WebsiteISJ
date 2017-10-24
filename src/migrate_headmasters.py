import os
import unicodedata
import string
from random import *

import django
from openpyxl import load_workbook, Workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from authentication.models import ExtendedUser
from school.models import School

wb = load_workbook('retea.xlsx')
to_write = Workbook()
ws1 = to_write.active
ws1.title = "Users"
counter_users = 0


def generate_password():
    characters = string.ascii_letters + string.punctuation + string.digits
    password = "".join(choice(characters) for x in range(randint(12, 16)))
    return password


for sheet in wb.get_sheet_names():
    for row in range(2, wb[sheet].max_row):
        school_name_file_row = 'B' + str(row)
        headmaster_name_file_row = 'D' + str(row)
        school_name = wb[sheet][school_name_file_row].value
        headmaster_name = wb[sheet][headmaster_name_file_row].value
        headmaster_splitted = headmaster_name.split(" ")
        username = (headmaster_splitted[0] + "_" +
                    headmaster_splitted[len(headmaster_splitted) - 1]).lower()
        username = unicodedata.normalize(
                "NFKD", username).encode("ascii", "ignore")
        first_name = username.split("_")[1].title()
        last_name = username.split("_")[0].title()
        try:
            school = School.objects.get(name=school_name)
            user = ExtendedUser.objects.create(username=username,
                                               first_name=first_name,
                                               last_name=last_name,
                                               status=1,
                                               school=school)
            counter_users = counter_users + 1
            password = generate_password()
            count_string = str(counter_users)
            ws1['A' + count_string] = username
            ws1['B' + count_string] = password
            user.set_password(password)
            user.save()
        except School.DoesNotExist:
            print username
to_write.save('users.xlsx')
